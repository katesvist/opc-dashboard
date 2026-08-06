from __future__ import annotations

import csv
import io
import re
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


MAX_FILE_SIZE = 10 * 1024 * 1024
MAX_ROWS = 20_000
SHEET_NAME = "Привязки"
HEADER_ROW = 4
DATA_START_ROW = HEADER_ROW + 1

COLUMNS: list[tuple[str, str, int, bool]] = [
    ("endpoint_id", "Endpoint\n(endpoint_id)", 24, False),
    ("node_id", "Node ID *\n(node_id)", 48, True),
    ("parameter_code", "Код параметра *\n(parameter_code)", 30, True),
    ("parameter_name", "Описание параметра\n(parameter_name)", 34, False),
    ("enabled", "Активна\n(enabled)", 13, False),
    ("acquisition_mode", "Режим\n(acquisition_mode)", 18, False),
    ("read_enabled", "Чтение\n(read_enabled)", 13, False),
    ("write_enabled", "Запись\n(write_enabled)", 13, False),
    ("sampling_interval_ms", "Подписка, мс\n(sampling_interval_ms)", 18, False),
    ("polling_interval_seconds", "Опрос, с\n(polling_interval_seconds)", 18, False),
    ("group_path", "Группа\n(group_path)", 28, False),
    ("browse_name", "Browse name\n(browse_name)", 24, False),
    ("display_name", "Название ноды\n(display_name)", 26, False),
    ("expected_type", "Тип\n(expected_type)", 16, False),
    ("value_shape", "Форма значения\n(value_shape)", 18, False),
    ("unit", "Ед. изм.\n(unit)", 16, False),
    ("tags", "Теги\n(tags)", 24, False),
]

KEYS = [column[0] for column in COLUMNS]
REQUIRED_KEYS = {"node_id", "parameter_code"}

ALIASES: dict[str, set[str]] = {
    "endpoint_id": {"endpoint", "endpointid", "источник", "кодисточника"},
    "node_id": {"nodeid", "node", "нода", "идноды", "идентификаторноды"},
    "parameter_code": {"кодпараметра", "параметр", "parameter", "parametercode", "paramcode"},
    "parameter_name": {"описаниепараметра", "имяпараметра", "parametername"},
    "enabled": {"активна", "активно", "enabled"},
    "acquisition_mode": {"режим", "mode", "acquisitionmode"},
    "read_enabled": {"чтение", "read", "readenabled"},
    "write_enabled": {"запись", "write", "writeenabled"},
    "sampling_interval_ms": {"подпискамс", "samplinginterval", "samplingintervalms"},
    "polling_interval_seconds": {"опросс", "pollinginterval", "pollingintervalseconds"},
    "group_path": {"группа", "путьгруппы", "grouppath"},
    "browse_name": {"browsename", "browse"},
    "display_name": {"названиеноды", "displayname"},
    "expected_type": {"тип", "expectedtype", "datatype"},
    "value_shape": {"формазначения", "valueshape", "shape"},
    "unit": {"едизм", "единицаизмерения", "unit", "uom"},
    "tags": {"теги", "tags"},
}

TRUE_VALUES = {"1", "true", "yes", "y", "да", "д", "вкл", "включено", "активна"}
FALSE_VALUES = {"0", "false", "no", "n", "нет", "н", "выкл", "выключено", "неактивна"}
MODE_VALUES = {
    "subscription": "subscription",
    "subscribe": "subscription",
    "sub": "subscription",
    "подписка": "subscription",
    "polling": "polling",
    "poll": "polling",
    "опрос": "polling",
}
TYPE_VALUES = {"bool", "int", "float", "str", "char", "datetime"}
SHAPE_VALUES = {"scalar", "array", "object"}


class BindingTableError(ValueError):
    pass


def export_bindings_xlsx(nodes: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.sheet_view.showGridLines = False
    instructions = workbook.create_sheet("Инструкция")
    instructions.sheet_view.showGridLines = False

    last_column = get_column_letter(len(COLUMNS))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = "Привязки OPC UA нод к параметрам"
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F5F8B")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30

    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = (
        "* обязательные поля. Пустые необязательные поля будут заполнены безопасными значениями. "
        "Импорт добавляет только новые привязки в черновик и не сохраняет их автоматически."
    )
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet["A2"].fill = PatternFill("solid", fgColor="EAF2F8")
    sheet.row_dimensions[2].height = 42

    required_fill = PatternFill("solid", fgColor="FFE3A3")
    optional_fill = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(style="thin", color="B9C1B7")
    for index, (_, title, width, required) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(HEADER_ROW, index, title)
        cell.font = Font(bold=True, color="1D2521")
        cell.fill = required_fill if required else optional_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=thin)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[HEADER_ROW].height = 42
    sheet.freeze_panes = f"A{DATA_START_ROW}"

    data_rows = [_node_to_row(node) for node in nodes]
    if not data_rows:
        data_rows = [[None] * len(COLUMNS)]
    for row in data_rows:
        sheet.append(row)

    last_data_row = HEADER_ROW + len(data_rows)
    table = Table(displayName="OpcUaBindings", ref=f"A{HEADER_ROW}:{last_column}{last_data_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    _add_validations(sheet)

    instruction_rows = [
        ["Как заполнять", "Значение"],
        ["Endpoint *", "ID источника из дашборда, например remote-opc-server."],
        ["Node ID *", "Точный OPC UA NodeId, например ns=3;s=\"DB\".\"Temperature\"."],
        ["Код параметра *", "Код name из Справочника параметров. Регистр и лишние пробелы исправляются при импорте."],
        ["Активна / Чтение / Запись", "Допустимы: Да/Нет, true/false, 1/0. По умолчанию: Да/Да/Нет."],
        ["Режим", "subscription (подписка) или polling (опрос). По умолчанию subscription."],
        ["Группа", "Путь через /, например _DB_FOR_TEST / ARRAY."],
        ["Теги", "Несколько тегов через запятую."],
        ["Конфликты", "Уже существующие NodeId и уже привязанные параметры не перезаписываются, а пропускаются с пояснением."],
    ]
    for row in instruction_rows:
        instructions.append(row)
    instructions.column_dimensions["A"].width = 30
    instructions.column_dimensions["B"].width = 100
    instructions.freeze_panes = "A2"
    instructions["A1"].font = instructions["B1"].font = Font(bold=True, color="FFFFFF")
    instructions["A1"].fill = instructions["B1"].fill = PatternFill("solid", fgColor="1F5F8B")
    for row in instructions.iter_rows(min_row=2, max_row=len(instruction_rows), min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    instructions.sheet_properties.pageSetUpPr.fitToPage = True

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def import_bindings_table(filename: str, content: bytes) -> dict[str, Any]:
    if not content:
        raise BindingTableError("Файл пуст.")
    if len(content) > MAX_FILE_SIZE:
        raise BindingTableError("Файл больше 10 МБ.")

    suffix = Path(filename or "").suffix.lower()
    if suffix == ".xlsx":
        raw_rows = _read_xlsx(content)
    elif suffix in {".csv", ".tsv", ".txt"}:
        raw_rows = _read_delimited(content, suffix)
    else:
        raise BindingTableError("Поддерживаются файлы .xlsx, .csv и .tsv.")

    return _normalize_rows(raw_rows)


def _read_xlsx(content: bytes) -> list[list[Any]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    except Exception as exc:
        raise BindingTableError(f"Не удалось прочитать XLSX: {exc}") from exc
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows: list[list[Any]] = []
    for index, row in enumerate(sheet.iter_rows(values_only=False), start=1):
        if index > MAX_ROWS + 30:
            raise BindingTableError(f"В таблице больше {MAX_ROWS} строк.")
        values = []
        for cell in row:
            if cell.data_type == "f":
                raise BindingTableError(f"Формулы не поддерживаются: ячейка {cell.coordinate}.")
            values.append(cell.value)
        rows.append(values)
    workbook.close()
    return rows


def _read_delimited(content: bytes, suffix: str) -> list[list[Any]]:
    text = _decode_text(content)
    delimiter = "\t" if suffix == ".tsv" else _detect_delimiter(text)
    rows = [list(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]
    if delimiter == ";":
        _repair_unquoted_node_ids(rows)
    return rows


def _detect_delimiter(text: str) -> str:
    lines = [line for line in text.splitlines()[:10] if line.strip()]
    if not lines:
        return ";"
    scores = {
        delimiter: sum(line.count(delimiter) for line in lines)
        for delimiter in (";", "\t", ",")
    }
    # При равенстве предпочитаем точку с запятой: запятая часто встречается
    # внутри русских заголовков и десятичных чисел.
    return max(scores, key=lambda delimiter: (scores[delimiter], delimiter == ";", delimiter == "\t"))


def _repair_unquoted_node_ids(rows: list[list[Any]]) -> None:
    """Склеивает OPC UA NodeId, если его `;` забыли заключить в кавычки CSV."""
    for header_index, header in enumerate(rows[:20]):
        keys = [_header_key(value)[0] for value in header]
        if "node_id" not in keys or "parameter_code" not in keys:
            continue
        node_column = keys.index("node_id")
        expected_width = len(header)
        for index in range(header_index + 1, len(rows)):
            row = rows[index]
            extra_cells = len(row) - expected_width
            if extra_cells <= 0:
                continue
            merged_node_id = ";".join(str(value) for value in row[node_column : node_column + extra_cells + 1])
            rows[index] = [
                *row[:node_column],
                merged_node_id,
                *row[node_column + extra_cells + 1 :],
            ]
        return


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise BindingTableError("Текстовый файл должен быть в UTF-8 или Windows-1251.")


def _normalize_rows(raw_rows: list[list[Any]]) -> dict[str, Any]:
    header_index, mapping, header_warnings = _find_header(raw_rows)
    rows: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []

    for source_index, raw in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
        if not any(_clean_text(value) for value in raw):
            continue
        if len(rows) >= MAX_ROWS:
            raise BindingTableError(f"В таблице больше {MAX_ROWS} строк данных.")

        item = {key: raw[column] if column < len(raw) else None for key, column in mapping.items()}
        normalized, row_issues = _normalize_item(item, source_index)
        if row_issues:
            issues.extend(row_issues)
        else:
            rows.append(normalized)

    return {
        "rows": rows,
        "issues": issues,
        "warnings": header_warnings,
        "source_rows": max(0, len(raw_rows) - header_index - 1),
        "accepted_rows": len(rows),
    }


def _find_header(raw_rows: list[list[Any]]) -> tuple[int, dict[str, int], list[str]]:
    for row_index, row in enumerate(raw_rows[:20]):
        mapping: dict[str, int] = {}
        warnings: list[str] = []
        for column_index, value in enumerate(row):
            key, fuzzy = _header_key(value)
            if key and key not in mapping:
                mapping[key] = column_index
                if fuzzy:
                    warnings.append(f"Заголовок «{_clean_text(value)}» распознан как {key}.")
        if REQUIRED_KEYS.issubset(mapping):
            return row_index, mapping, warnings
    raise BindingTableError("Не найдена строка заголовков с Node ID и кодом параметра.")


def _header_key(value: Any) -> tuple[str | None, bool]:
    normalized = _normalize_header(value)
    if not normalized:
        return None, False
    for key in KEYS:
        candidates = {_normalize_header(key), *ALIASES.get(key, set())}
        if normalized in candidates:
            return key, False
    scores: list[tuple[float, str]] = []
    for key in KEYS:
        candidates = {_normalize_header(key), *ALIASES.get(key, set())}
        score = max(SequenceMatcher(None, normalized, candidate).ratio() for candidate in candidates)
        scores.append((score, key))
    scores.sort(reverse=True)
    if scores[0][0] >= 0.88 and (len(scores) == 1 or scores[0][0] - scores[1][0] >= 0.08):
        return scores[0][1], True
    return None, False


def _normalize_item(item: dict[str, Any], row_number: int) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    issues: list[dict[str, Any]] = []
    normalized: dict[str, Any] = {key: _clean_text(item.get(key)) or None for key in KEYS}
    normalized["source_row"] = row_number
    for key in ("endpoint_id", "node_id", "parameter_code"):
        normalized[key] = _clean_identifier(item.get(key)) or None

    for required in REQUIRED_KEYS:
        if not normalized.get(required):
            issues.append({"row": row_number, "field": required, "message": "Обязательное поле не заполнено."})

    for key, default in (("enabled", True), ("read_enabled", True), ("write_enabled", False)):
        parsed = _parse_bool(item.get(key), default)
        if parsed is None:
            issues.append({"row": row_number, "field": key, "message": f"Неизвестное значение: {_clean_text(item.get(key))}."})
        normalized[key] = parsed

    mode_raw = (_clean_text(item.get("acquisition_mode")) or "subscription").lower()
    normalized["acquisition_mode"] = MODE_VALUES.get(mode_raw)
    if normalized["acquisition_mode"] is None:
        issues.append({"row": row_number, "field": "acquisition_mode", "message": f"Допустимы subscription или polling, получено: {mode_raw}."})

    normalized["sampling_interval_ms"] = _parse_number(item.get("sampling_interval_ms"), 1000, integer=True)
    normalized["polling_interval_seconds"] = _parse_number(item.get("polling_interval_seconds"), 5.0)
    if normalized["sampling_interval_ms"] is None or normalized["sampling_interval_ms"] <= 0:
        issues.append({"row": row_number, "field": "sampling_interval_ms", "message": "Интервал должен быть положительным целым числом."})
    if normalized["polling_interval_seconds"] is None or normalized["polling_interval_seconds"] <= 0:
        issues.append({"row": row_number, "field": "polling_interval_seconds", "message": "Интервал должен быть положительным числом."})

    expected_type = (_clean_text(item.get("expected_type")) or "").lower()
    normalized["expected_type"] = expected_type if expected_type in TYPE_VALUES else None
    if expected_type and normalized["expected_type"] is None:
        issues.append({"row": row_number, "field": "expected_type", "message": f"Неизвестный тип: {expected_type}."})

    value_shape = (_clean_text(item.get("value_shape")) or "scalar").lower()
    normalized["value_shape"] = value_shape if value_shape in SHAPE_VALUES else None
    if normalized["value_shape"] is None:
        issues.append({"row": row_number, "field": "value_shape", "message": f"Допустимы scalar, array или object, получено: {value_shape}."})

    normalized["group_path"] = _split_values(item.get("group_path"), separators=r"[/\\>|]+")
    normalized["tags"] = _split_values(item.get("tags"), separators=r"[,;]+")
    return normalized, issues


def _node_to_row(node: dict[str, Any]) -> list[Any]:
    values: dict[str, Any] = dict(node)
    values["enabled"] = _yes_no(node.get("enabled", True))
    values["read_enabled"] = _yes_no(node.get("read_enabled", True))
    values["write_enabled"] = _yes_no(node.get("write_enabled", False))
    group_path = node.get("group_path")
    tags = node.get("tags")
    values["group_path"] = " / ".join(group_path) if isinstance(group_path, list) else ""
    values["tags"] = ", ".join(tags) if isinstance(tags, list) else ""
    return [_safe_excel_value(values.get(key)) for key in KEYS]


def _add_validations(sheet: Any) -> None:
    validations = {
        "E": '"Да,Нет"',
        "F": '"subscription,polling"',
        "G": '"Да,Нет"',
        "H": '"Да,Нет"',
        "N": '"bool,int,float,str,char,datetime"',
        "O": '"scalar,array,object"',
    }
    for column, formula in validations.items():
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "Выберите значение из списка."
        validation.errorTitle = "Неверное значение"
        validation.prompt = "Можно выбрать значение из списка."
        validation.promptTitle = "Формат поля"
        sheet.add_data_validation(validation)
        validation.add(f"{column}{DATA_START_ROW}:{column}{MAX_ROWS + DATA_START_ROW}")


def _normalize_header(value: Any) -> str:
    text = _clean_text(value).lower()
    technical = re.findall(r"\(([a-z_]+)\)", text)
    if technical:
        return re.sub(r"[^a-zа-яё0-9]", "", technical[-1])
    return re.sub(r"[^a-zа-яё0-9]", "", text)


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()


def _clean_identifier(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip()
    if len(text) > 1 and text[0] == "'" and text[1] in "=+-@":
        return text[1:]
    return text


def _parse_bool(value: Any, default: bool) -> bool | None:
    text = _clean_text(value).lower()
    if not text:
        return default
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return None


def _parse_number(value: Any, default: float | int, *, integer: bool = False) -> float | int | None:
    text = _clean_text(value)
    if not text:
        return default
    try:
        number = float(text.replace(" ", "").replace(",", "."))
    except ValueError:
        return None
    if integer:
        return int(number) if number.is_integer() else None
    return number


def _split_values(value: Any, *, separators: str) -> list[str]:
    return [part.strip() for part in re.split(separators, _clean_text(value)) if part.strip()]


def _yes_no(value: Any) -> str:
    return "Да" if bool(value) else "Нет"


def _safe_excel_value(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value
